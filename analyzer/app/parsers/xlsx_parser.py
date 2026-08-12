"""XLSX extraction.

Read-only, values-only, streamed. A Document Control workbook - a form log, a
register - can hold hundreds of thousands of cells, and loading one with
formulas and styling would cost far more memory than the text is worth
(CLAUDE.md 35.12).

The worksheet name becomes the section, which is the natural unit an operator
would use to describe where a translation is missing.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.config import get_settings
from app.parsers.base import (
    DocumentParser,
    ExtractedDocument,
    ParserError,
    SegmentKind,
    TextSegment,
)


class XlsxParser(DocumentParser):
    extensions = ("xlsx",)
    name = "xlsx"

    def parse(self, path: Path) -> ExtractedDocument:
        settings = get_settings()

        try:
            workbook = load_workbook(
                filename=str(path),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except InvalidFileException as exc:
            raise ParserError("The spreadsheet is not a valid XLSX file.") from exc
        except Exception as exc:
            raise ParserError("The spreadsheet could not be opened. It may be corrupt.") from exc

        result = ExtractedDocument(parser=self.name)
        cells_read = 0

        try:
            for worksheet in workbook.worksheets:
                if cells_read >= settings.max_xlsx_cells:
                    result.truncated = True
                    result.notes.append(
                        f"Extraction stopped at {settings.max_xlsx_cells} cells; "
                        f"sheet '{worksheet.title}' onwards was not analysed."
                    )
                    break

                for row in worksheet.iter_rows(values_only=True):
                    for value in row:
                        cells_read += 1

                        if cells_read >= settings.max_xlsx_cells:
                            result.truncated = True
                            break

                        # Only strings carry language. Numbers, dates and
                        # booleans are data, and feeding them to a language
                        # detector would be noise.
                        if isinstance(value, str):
                            text = value.strip()
                            if text:
                                result.segments.append(
                                    TextSegment(
                                        text=text,
                                        kind=SegmentKind.SHEET_CELL,
                                        section=worksheet.title,
                                    )
                                )

                    if result.truncated:
                        break
        except Exception as exc:
            raise ParserError("The spreadsheet could not be read to the end.") from exc
        finally:
            workbook.close()

        return result
